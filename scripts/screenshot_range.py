#!/usr/bin/env python3
"""Screenshot xlsx cell range -> clean HTML table.

Approach: build HTML table cell-by-cell from openpyxl.
- Skips hidden rows/cols 
- Merged cells -> colspan/rowspan, child cells omitted
- Bold cells -> <b> tag
- Clean CSS, no Excel cloning

Usage: screenshot_range.py <xlsx> <sheet> <range> [out.png]
"""
import os, sys, re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from playwright.sync_api import sync_playwright

def parse_range(s):
    m = re.match(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', s.upper())
    if not m: raise ValueError(s)
    return (column_index_from_string(m[1]), int(m[2]),
            column_index_from_string(m[3]), int(m[4]))

def theme_to_rgb(theme_idx, tint=0):
    """Map Excel theme color to RGB. tint: 0=full, +lighter, -darker"""
    themes = {0:'FFFFFF',1:'000000',2:'F2F2F2',3:'7F7F7F',
              4:'4472C4',5:'ED7D31',6:'A5A5A5',7:'FFC000',8:'5B9BD5',9:'70AD47'}
    base = themes.get(theme_idx, 'FFFFFF')
    if tint == 0: return base
    mix = 'FFFFFF' if tint > 0 else '000000'
    t = abs(tint)
    r = int(int(base[0:2],16)*(1-t) + int(mix[0:2],16)*t)
    g = int(int(base[2:4],16)*(1-t) + int(mix[2:4],16)*t)
    b = int(int(base[4:6],16)*(1-t) + int(mix[4:6],16)*t)
    return f'{r:02X}{g:02X}{b:02X}'

def get_color_hex(color_obj):
    """Get #RRGGBB from openpyxl Color (handles theme + rgb)"""
    if not color_obj: return None
    try:
        if color_obj.type == 'theme' and color_obj.theme is not None:
            return theme_to_rgb(color_obj.theme, color_obj.tint or 0)
        if color_obj.type == 'rgb':
            r = str(color_obj.rgb)
            if r and r != '00000000' and len(r) >= 6:
                return r[-6:]
    except: pass
    return None

def fmt_val(val, number_format=None):
    """Format cell value for display."""
    import datetime
    if val is None: return ''
    if isinstance(val, datetime.datetime):
        nf = number_format or ''
        # Map Excel date formats to strftime
        fmts = {
            'mm-dd-yy': '%m-%d-%y', 'yyyy-mm-dd': '%Y-%m-%d', 'yyyy/mm/dd': '%Y/%m/%d',
            'm/d/yyyy': '%m/%d/%Y', 'm/d/yy': '%m/%d/%y', 'd/m/yyyy': '%d/%m/%Y',
            'yyyy\\-mm\\-dd': '%Y-%m-%d', 'mm\\-dd\\-yy': '%m-%d-%y',
        }
        for xl_fmt, py_fmt in fmts.items():
            if xl_fmt in nf:
                return val.strftime(py_fmt)
        return val.strftime('%Y-%m-%d')
    if isinstance(val, datetime.timedelta):
        nf = number_format or ''
        total_sec = val.total_seconds()
        if 'hh:mm:ss' in nf or 'h:mm:ss' in nf:
            h = int(total_sec // 3600)
            m = int((total_sec % 3600) // 60)
            s = int(total_sec % 60)
            return f'{h:02d}:{m:02d}:{s:02d}'
        return str(val)
    if isinstance(val, float):
        nf = number_format or ''
        if '0.00%' in nf or '0%' in nf or '0.0%' in nf:
            return f'{val*100:.2f}%'
        if val == int(val):
            return f'{int(val):,}'
        return f'{val:,.2f}'
    return str(val)

def main():
    if len(sys.argv) < 4:
        print("Usage: screenshot_range <xlsx> <sheet> <range> [out.png]"); sys.exit(1)
    xlsx, sheet, rng = sys.argv[1], sys.argv[2], sys.argv[3]
    out = sys.argv[4] if len(sys.argv) > 4 else '/tmp/_ss.png'
    c1, r1, c2, r2 = parse_range(rng)

    wb = load_workbook(xlsx, data_only=True)
    ws = wb[sheet]

    # Hidden
    H = {r for r in range(r1, r2+1) if ws.row_dimensions.get(r) and ws.row_dimensions[r].hidden}
    HC = {c for c in range(c1, c2+1) if ws.column_dimensions.get(get_column_letter(c)) and ws.column_dimensions[get_column_letter(c)].hidden}

    # Merge map: root(r,c) -> (rowspan,colspan), children set
    # CRITICAL: rowspan counts only VISIBLE rows within the merge range
    MR = {}; MC = set()
    for mr in ws.merged_cells.ranges:
        mc1, mr1, mc2, mr2_ = parse_range(str(mr))
        if mr1 in H or mc1 in HC: continue
        # Count visible rows and cols within merge
        vis_rows = sum(1 for rr in range(mr1, mr2_ + 1) if rr not in H)
        vis_cols = sum(1 for cc in range(mc1, mc2 + 1) if cc not in HC)
        MR[(mr1, mc1)] = (vis_rows, vis_cols)
        for rr in range(mr1, mr2_ + 1):
            for cc in range(mc1, mc2 + 1):
                if (rr, cc) != (mr1, mc1):
                    MC.add((rr, cc))

    # Column widths
    CW = []
    for c in range(c1, c2+1):
        if c in HC: continue
        cd = ws.column_dimensions.get(get_column_letter(c))
        CW.append(max(int(cd.width * 8) if cd and cd.width else 80, 50))

    # Build HTML: one pass, cell by cell
    html = []
    alt_css = 'tr:nth-child(even) td{background:#f2f2f2}'  # fallback, cell styles override

    html.append(f'''<!DOCTYPE html><html><head><meta charset="utf-8"><style>
*{{margin:0;padding:0}}body{{background:#fff;font-family:"Microsoft YaHei","SimHei",sans-serif;font-size:12px}}
table{{border-collapse:collapse}}
td{{padding:4px 8px;border:1px solid #d0d0d0;white-space:nowrap}}
{alt_css}
</style></head><body><table><colgroup>''')
    for w in CW:
        html.append(f'<col style="width:{w}px">')
    html.append('</colgroup>')

    for r in range(r1, r2+1):
        if r in H: continue
        html.append('<tr>')
        for c in range(c1, c2+1):
            if c in HC: continue
            if (r, c) in MC: continue  # skip merge children
            cell = ws.cell(r, c)
            val = fmt_val(cell.value, cell.number_format)
            is_bold = cell.font and cell.font.bold
            
            # Read background color
            bg = ''
            hex_bg = get_color_hex(cell.fill.fgColor) if cell.fill and cell.fill.patternType else None
            if hex_bg:
                bg = f'background-color:#{hex_bg};'
            
            # Read font color
            fc = ''
            hex_fc = get_color_hex(cell.font.color) if cell.font else None
            if hex_fc:
                fc = f'color:#{hex_fc};'
            
            # Read alignment
            al = ''
            if cell.alignment:
                h = str(cell.alignment.horizontal) if cell.alignment.horizontal else ''
                v = str(cell.alignment.vertical) if cell.alignment.vertical else ''
                if h in ('center','left','right','justify'):
                    al += f'text-align:{h};'
                if v in ('top','center','bottom'):
                    al += f'vertical-align:{v};'
            
            style = bg + fc + al
            tag = '<td' + (f' style="{style}"' if style else '')
            if (r, c) in MR:
                rs, cs = MR[(r, c)]
                tag += f' colspan="{cs}" rowspan="{rs}"'
            if is_bold:
                tag += '><b>' + val + '</b></td>'
            else:
                tag += '>' + val + '</td>'
            html.append(tag)
        html.append('</tr>')
    html.append('</table></body></html>')
    full = '\n'.join(html)

    hpath = '/var/www/html/tmp/_ss_' + str(os.getpid()) + '.html'
    with open(hpath, 'w', encoding='utf-8') as f:
        f.write(full)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True, args=['--no-sandbox'])
        page = browser.new_page(viewport={'width':2560,'height':1600})
        page.goto(f'file://{hpath}')
        page.wait_for_timeout(400)
        box = page.locator('table').bounding_box()
        if box:
            page.screenshot(path=out, clip={'x':box['x'],'y':box['y'],'width':box['width'],'height':box['height']}, full_page=True)
        else:
            page.screenshot(path=out, full_page=True)
        browser.close()
    print(f'OK: {out} ({os.path.getsize(out)}B)')

if __name__ == '__main__':
    main()
