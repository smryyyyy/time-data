#!/usr/bin/env python3
"""
合并数据：openpyxl 仅修改 今日/昨日 sheet 的指定列，保护所有公式和图表
"""
import sys, os
from openpyxl import load_workbook

def main():
    if len(sys.argv) < 7:
        print("用法: merge_oa.py <date> <prefix> <start_col> <end_col> <data_dir> <template_file>")
        sys.exit(1)

    date_str  = sys.argv[1]
    prefix    = sys.argv[2]
    start_col = sys.argv[3]  # e.g. 'A'
    end_col   = sys.argv[4]  # e.g. 'K'
    data_dir  = sys.argv[5]
    tmpl_file = sys.argv[6]

    # 计算昨日
    from datetime import datetime, timedelta
    d = datetime.strptime(date_str, '%Y-%m-%d')
    yesterday = (d - timedelta(days=1)).strftime('%Y-%m-%d')

    today_src = os.path.join(data_dir, date_str, f'{prefix}({date_str}).xlsx')
    yest_src  = os.path.join(data_dir, yesterday, f'{prefix}({yesterday}).xlsx')

    # 验证文件
    for f in [tmpl_file, today_src, yest_src]:
        if not os.path.exists(f):
            print(f'ERROR: 文件不存在: {f}')
            sys.exit(1)

    # 列范围 → 列索引 (A=1, K=11)
    col_start = ord(start_col.upper()) - ord('A') + 1
    col_end   = ord(end_col.upper()) - ord('A') + 1

    # 读取源数据
    def read_source(path):
        wb = load_workbook(path, data_only=True)
        sheet = wb['广告主_报告'] if '广告主_报告' in wb.sheetnames else wb.active
        data = []
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=col_start, max_col=col_end, values_only=True):
            data.append(list(row))
        wb.close()
        return data

    today_data = read_source(today_src)
    yest_data  = read_source(yest_src)

    # 打开模板，写入数据
    wb = load_workbook(tmpl_file)
    jinri  = wb['今日']
    zuori  = wb['昨日']

    # 清空并写入今日
    for ri, row_data in enumerate(today_data, start=1):
        for ci, val in enumerate(row_data):
            jinri.cell(row=ri, column=col_start + ci).value = val

    # 清空并写入昨日
    for ri, row_data in enumerate(yest_data, start=1):
        for ci, val in enumerate(row_data):
            zuori.cell(row=ri, column=col_start + ci).value = val

    # 直接覆盖保存（保护所有其他sheet的公式和图表）
    wb.save(tmpl_file)
    wb.close()

    print(f'OK: 今日 {len(today_data)} 行, 昨日 {len(yest_data)} 行 → {tmpl_file}')

if __name__ == '__main__':
    main()
