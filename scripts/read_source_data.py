#!/usr/bin/env python3
"""Read source data sheet from xlsx using openpyxl (low memory).
Usage: read_source_data.py <xlsx_file> <columns_csv>
Outputs JSON array of row data to stdout.
"""
import sys, json
from openpyxl import load_workbook

if len(sys.argv) < 3:
    print("Usage: read_source_data.py <xlsx> <A,B,C>", file=sys.stderr)
    sys.exit(1)

xlsx = sys.argv[1]
cols = sys.argv[2].split(',')

# read_only=True keeps memory usage low
wb = load_workbook(xlsx, read_only=True, data_only=True)
ws = wb['广告主_报告']

result = []
for row in ws.iter_rows(min_row=1, values_only=True):
    row_data = {}
    for i, col in enumerate(cols):
        if i < len(row):
            val = row[i]
            row_data[col] = val
        else:
            row_data[col] = None
    result.append(row_data)

wb.close()
print(json.dumps(result, ensure_ascii=False, default=str))
